from collections import Counter
from openpyxl import workbook

def count_barrage():
    
    #获得弹幕列表
    
    file = open('弹幕1.txt', 'r',encoding = 'utf-8')
    barrage_data = file.read()
    barrage_list = barrage_data.split('\n')
    
    #统计列表中各弹幕的数量并排序
    
    counter = Counter(barrage_list)
    counter_result = counter.most_common()
    
    #输出数量前20的弹幕
    
    for barrage,count in counter_result[:20]:
        print(barrage,count)
        
    #存入barrage_count.xlsx
        
    data = workbook.Workbook()
    sheet = data.worksheets[0]
    sheet.cell(1,1).value = "弹幕"
    sheet.cell(1,2).value = "数量"
    row=2
    for barrage,count in counter_result:
        sheet.cell(row,1).value=barrage
        sheet.cell(row,2).value=count
        row=row+1
    data.save('barrage_count.xlsx')
    file.close()
    
if __name__ == "__main__":
    count_barrage()